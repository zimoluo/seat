from src.constants import *
from src.reader import Config
import json
from openpyxl import load_workbook
import time
from src.seat import Util

class FileExtension:
    def __init__(self, folder: str):
        self._folder = folder
        with open(f'./save/{folder}/config.json', mode='r', encoding='utf-8') as config:
            self._config = Config(json.loads(config.read()))
        self._seats = self._config.seats
    
    def getTxt(self) -> None:
        path = f'./save/{self._folder}/{self._folder} seats.txt'
        with open(path, mode='w', encoding='utf-8') as text:
            text.truncate()
            text.write(str(self._seats))
    
    def getStr(self) -> str:
        return str(self._seats)
    
    def getExcelForm(self) -> None:
        path = f'./save/{self._folder}/{self._folder} seats.xlsx'
        path_template = f'./save/{self._folder}/template.xlsx'
    
        date = [time.strftime('%d', time.localtime()).strip(), time.strftime('%b', time.localtime()).strip(), time.strftime('%Y', time.localtime()).strip()]
        wb = load_workbook(path_template)
        sheet = wb.active

        start = FileExtension.findStart(sheet)

        for rowSeat in range(len(self._seats)):
            rowForm = rowSeat + start['row']
            for colSeat in range(len(self._seats[0]) - 1, -1, -1):
                colForm = len(self._seats[0]) - 1 - colSeat + start['col']

                student = self._seats[rowSeat, colSeat]
                
                if student.sitter == EMPTY:
                    continue

                sheet.cell(row=rowForm, column=colForm).value = student.sitter
        
        try:
            dateline = sheet.cell(row=start['date'][0], column=start['date'][1]).value

            dateline = Util.lineWordReplace(dateline, 'DAY', date[0])
            dateline = Util.lineWordReplace(dateline, 'MONTH', date[1])
            dateline = Util.lineWordReplace(dateline, 'YEAR', date[2])
            sheet.cell(row=start['date'][0], column=start['date'][1]).value = dateline
        except KeyError: ...

        wb.save(path)

    @staticmethod
    def findStart(sheet) -> dict:
        start = {'row': 1, 'col': 1}

        for row in sheet.rows:
            for cell in row:
                if cell.value is None:
                    continue
                
                if cell.value[0] == '$':
                    if 'ROW' in cell.value:
                        start['row'] = cell.row

                    if 'COL' in cell.value:
                        start['col'] = cell.column

                    cell.value = None
                    continue

                if r'%' == cell.value[0]:
                    start['date'] = cell.row, cell.column
                    cell.value = cell.value[1:]

        return start